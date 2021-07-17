import telebot
import pickle
import openpyxl
import random

token="15588458:AAGk0B-nEy1PnkA3PSJTxubnPAmc43zwyoY"
bot = telebot.TeleBot(token)
def crypt(mem):
    with open('members.txt', 'wb') as out:
        pickle.dump(mem, out)
def decrypt():
    with open('members.txt','rb') as inp:
        mem = pickle.load(inp)
    return mem
# members={id:[Время,[Продукты],позиция,столбец,позиция_для_кнопок, позиция_для_категорий]}
def take_from_cordinate(id):
    wb = openpyxl.load_workbook('Prodvkt.xlsx')
    sheet = wb.get_sheet_by_name('Products')
    members=decrypt()
    mass_with_id=members[id][1]
    crypt(members)
    name_rez=[]
    for i in mass_with_id:
        name_rez.append(sheet[str(i)].value)
    return name_rez
def take_dinner(id):
    bot.send_chat_action(id, 'typing')
    members=decrypt()
    f=members[id][0]
    crypt(members)
    if f=="1":
        e='Zavtrak'
    elif f=="2":
        e='Obed'
    else:
        e='Vzin'
    wb = openpyxl.load_workbook('Prodvkt.xlsx')
    sheet = wb.get_sheet_by_name(e)
    mass_for_rez_mass=[]
    y=-1
    for cellObj in sheet['A1':'T10']:
        if (cellObj[0].value)!=None:
            y+=1
            mass_for_rez_mass.append([])
        for cells in cellObj:
            if cells.value!=None:
                mass_for_rez_mass[y].append(str(cells.value).capitalize())
    return mass_for_rez_mass

def Choice(id,user_ch,system_ch):
    MASS=[]
    for i in system_ch:
        MASS.append(set(user_ch) & set(i))
    dlina=[]
    for i in MASS:
        dlina.append(len(i))
    return dlina.index(max(dlina))


def rez_function(id):
    bot.send_chat_action(id, 'typing')
    name_from_take=take_from_cordinate(id)
    mass_for_choice=take_dinner(id)
    number_of_str=Choice(id,name_from_take,mass_for_choice)
    wb = openpyxl.load_workbook('Prodvkt.xlsx')
    members = decrypt()
    f = members[id][0]
    crypt(members)
    if f == "1":
        e = 'Zavtrak'
    elif f == "2":
        e = 'Obed'
    else:
        e = 'Vzin'
    sheet = wb.get_sheet_by_name(e)
    bot.send_message(id,"Блюдо для вас называется: '"+sheet["U" + str(number_of_str + 1)].value+"'")
    bot.send_message(id, "Необходимые продукты для этого блюда: '" + sheet["V" + str(number_of_str + 1)].value + "'")

    if len(sheet["W" + str(number_of_str + 1)].value)>4000:
        fil=0
        k=sheet["W" + str(number_of_str + 1)].value
        while fil<=len(sheet["W" + str(number_of_str + 1)].value):
            efo=k[fil:fil+4000]
            bot.send_message(id,efo)
            fil+=4000
    else:
        bot.send_message(id,sheet["W" + str(number_of_str + 1)].value)
    user_markup = telebot.types.ReplyKeyboardMarkup(True)
    user_markup.row('Новый рецепт')
    bot.send_message(id, "Я надеючь, что мой рецепт помог провести время вкусно. Нажми на кнопку 'Новый рецепт' для указания других продуктов и получения нового рецепта.", reply_markup=user_markup)


def razb_na_str(id,rez_stolb):
    buttons_added = [[],[],[]]
    members=decrypt()
    six_rez=rez_stolb
    if len(rez_stolb)>6:
        six_rez=rez_stolb[members[id][4]:members[id][4]+6]
    e=0
    for i in six_rez:
        if len(buttons_added[e])<2:
            buttons_added[e].append(i.capitalize())
        else:
            e+=1
            buttons_added[e].append(i.capitalize())
    user_markup = telebot.types.ReplyKeyboardMarkup(True)
    user_markup.row(*buttons_added[0])
    user_markup.row(*buttons_added[1])
    user_markup.row(*buttons_added[2])
    if len(rez_stolb) > 6:
        if members[id][4]==0:
            buttons_to_next=[" ","Далее"]
        elif members[id][4]+6>=len(rez_stolb):
            buttons_to_next=["Назад"," "]
        else:
            buttons_to_next=["Назад","Далее"]
        user_markup.row(*buttons_to_next)
    user_markup.row("К категориям","Получить рецепт")
    bot.send_message(id, "Выбери продукты из этой категории.", reply_markup=user_markup)

def first_line_in_Excel(Need_type): # Считывает первую строку с категориями
    wb = openpyxl.load_workbook('Prodvkt.xlsx')
    sheet = wb.get_sheet_by_name('Products')
    mass_with_products = []
    list_with_products = {}
    first_row = (list(sheet.rows)[0])

    for j in first_row:
        if (j.value)!=None:
            mass_with_products.append(str(j.value).strip())
    for j in first_row:
        if (j.value) != None:
            list_with_products[str(j.value).strip()]=(j.coordinate[0])
    if Need_type=="Mass":
        return mass_with_products
    elif Need_type=="List":
        return  list_with_products

def list_prodvkt_in_Excel(Need_type,id, Name=None,stolb=None): # Считывает столбец
    wb = openpyxl.load_workbook('Prodvkt.xlsx')
    sheet = wb.get_sheet_by_name('Products')
    if stolb==None:
        stolb = first_line_in_Excel("List")[Name]
    column = sheet[stolb]
    members = decrypt()
    members[id][3] = stolb
    crypt(members)
    mass_column_rez = []
    list_column_rez = {}
    for i in range(len(column)):
        if column[i].value != None:
            list_column_rez[(str(column[i].value)).strip()]=column[i].coordinate
            mass_column_rez.append((str(column[i].value)).strip())
    if Need_type == "Mass":
        del mass_column_rez[0]
        return mass_column_rez
    if Need_type == "List":
        del list_column_rez[list(list_column_rez.keys())[0]]
        return  list_column_rez

def list_prodvkt(id,Name,stolb_name): #Выводит столбец из спискка продуктов
    if Name!=None:
        mass_with_products=list_prodvkt_in_Excel("Mass",id,Name)
    else:
        mass_with_products = list_prodvkt_in_Excel("Mass", id, None,stolb_name)
    razb_na_str(id,mass_with_products)

def prodvkt(id): # Выводит категории продуктов
    buttons_added = [[], [], []]
    mass_with_products=first_line_in_Excel("Mass")
    members = decrypt()
    six_rez = first_line_in_Excel("Mass")
    if len(mass_with_products) > 6:
        six_rez = mass_with_products[members[id][5]:members[id][5] + 6]
    e = 0
    for i in six_rez:
        if len(buttons_added[e]) < 2:
            buttons_added[e].append(i)
        else:
            e += 1
            buttons_added[e].append(i)
    user_markup = telebot.types.ReplyKeyboardMarkup(True)
    user_markup.row(*buttons_added[0])
    user_markup.row(*buttons_added[1])
    user_markup.row(*buttons_added[2])
    if len(mass_with_products) > 6:
        if members[id][5] == 0:
            buttons_to_next = [" ", "Следующие"]
        elif members[id][5] + 6 >= len(mass_with_products):
            buttons_to_next = ["Предыдущие", " "]
        else:
            buttons_to_next = ["Предыдущие", "Следующие"]
        user_markup.row(*buttons_to_next)
    user_markup.row("Обратно", "Получить рецепт")
    bot.send_message(id, "Выбери категорию продуктов.", reply_markup=user_markup)

def time_for_eat(id):
    user_markup = telebot.types.ReplyKeyboardMarkup(True)
    user_markup.row('Завтрак')
    user_markup.row('Обед')
    user_markup.row('Ужин')
    user_markup.row('Случайное блюдо')  # Доделать
    bot.send_message(id, "Выбери время своей трапезы.", reply_markup=user_markup)
@bot.message_handler(commands=['start'])
def handle_start(message):
    bot.send_message(message.chat.id, 'Привет. Я бот, который может помочь с выбором рецепта приготовления блюда из того, что ты укажешь.')
    time_for_eat(message.from_user.id)

@bot.message_handler(content_types=['text'])
def handle_message(message):
    if message.text.lower()== "завтрак": # Код 1
        members=decrypt()
        members[message.chat.id]=["1",[],"1","",0,0]
        crypt(members)
        prodvkt(message.chat.id)
    elif message.text.lower()== "обед": # Код 2
        members=decrypt()
        members[message.chat.id]=["2",[],"1","",0,0]
        crypt(members)
        prodvkt(message.chat.id)
    elif message.text.lower()== "ужин": # Код 3
        members=decrypt()
        members[message.chat.id]=["3",[],"1","",0,0]
        crypt(members)
        prodvkt(message.chat.id)
    elif (message.text.capitalize() in first_line_in_Excel("Mass")):# Выбор категории
        members=decrypt()
        if message.chat.id in members.keys():
            if members[message.chat.id][2]=="1":
                members[message.chat.id][2] = "2"
                crypt(members)
                list_prodvkt(message.chat.id,message.text.capitalize(),None)
        else:
            crypt(members)
    elif message.text.capitalize()=="Далее":
        members = decrypt()
        stolb_name = members[message.chat.id][3]
        mass_with_products = list_prodvkt_in_Excel("Mass", message.chat.id, None, stolb_name)
        if members[message.chat.id][4] + 6<len(mass_with_products):
            members[message.chat.id][4]=members[message.chat.id][4] + 6
        f=members[message.chat.id][3]
        crypt(members)
        list_prodvkt(message.chat.id,None, f)
    elif message.text.capitalize()=="Назад":
        members = decrypt()
        if members[message.chat.id][4] - 6 >= 0:
            members[message.chat.id][4] = members[message.chat.id][4] - 6
        f = members[message.chat.id][3]
        crypt(members)
        list_prodvkt(message.chat.id, None, f)
    elif message.text == "К категориям":
        bot.send_message(message.chat.id,"Хорошо.")
        members = decrypt()
        members[message.chat.id][4] = 0
        members[message.chat.id][2] = "1"
        crypt(members)
        prodvkt(message.chat.id)
    elif message.text=="Обратно":
        time_for_eat(message.chat.id)
    elif message.text == "Предыдущие":
        members = decrypt()
        if members[message.chat.id][5] - 6 >= 0:
            members[message.chat.id][5] = members[message.chat.id][5] - 6
        crypt(members)
        prodvkt(message.chat.id)
    elif message.text=="Следующие":
        members = decrypt()
        if members[message.chat.id][5] + 6 < len(first_line_in_Excel("Mass")):
            members[message.chat.id][5] = members[message.chat.id][5] + 6
        crypt(members)
        prodvkt(message.chat.id)
    elif message.text=="Получить рецепт":
        bot.send_message(message.chat.id,"Подбираю рецепт.")
        bot.send_chat_action(message.chat.id, 'typing')
        rez_function(message.chat.id)
    elif message.text == "Новый рецепт":
        time_for_eat(message.chat.id)
    elif message.text=="Случайное блюдо":
        f=random.randint(1,3)
        if f == "1":
            e = 'Zavtrak'
        elif f == "2":
            e = 'Obed'
        else:
            e = 'Vzin'
        wb = openpyxl.load_workbook('Prodvkt.xlsx')
        sheet = wb.get_sheet_by_name(e)
        i = "U" + str(random.randint(1, 10)) # Плохо работает, так как задан конец - 10
        while sheet[str(i)].value==None:
            i="U"+str(random.randint(1,10))
        bot.send_message(message.chat.id, "Блюдо для вас называется: '" + sheet[i].value + "'")
        bot.send_message(message.chat.id,
                         "Необходимые продукты для этого блюда: '" + sheet["v"+i[1]].value + "'")
        bot.send_message(message.chat.id, sheet["W"+i[1]].value)
    else: # Запись ячеек продуктов
        members = decrypt()
        if members[message.chat.id][2] == "2":
            if message.text.capitalize() in list_prodvkt_in_Excel("Mass",message.chat.id,None,members[message.chat.id][3]):
                list=list_prodvkt_in_Excel("List", message.chat.id, None, members[message.chat.id][3])
                if (list[message.text.capitalize()] in members[message.chat.id][1])==False:
                    members[message.chat.id][1].append(list[message.text.capitalize()])
                bot.send_message(message.chat.id,"Запомнил.")
                crypt(members)
        else:
            crypt(members)

bot.polling(none_stop=True)
